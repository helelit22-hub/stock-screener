#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
סורק מניות — NYSE + NASDAQ  (גרסת ענן / GitHub Actions)
מחפש מניות עם: שווי שוק > $100M, מחיר > SMA150, ATR(14) בין 3.5-7.0
שליחת מייל דרך SMTP (Gmail) במקום Apple Mail.
"""

import os
import sys
import json
import time
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta

import yfinance as yf
import pandas as pd
import numpy as np
import requests

# ─── קונפיגורציה ─────────────────────────────────────
MAX_RESULTS = 20         # להציג רק TOP 20 הגדולות לפי שווי שוק
HOT_PICK_DAYS = 3        # מניה שמופיעה X ימים ברצף = Hot Pick
HISTORY_FILE = 'history.json'
HISTORY_KEEP_DAYS = 30
VOL_SPIKE_RATIO = 3.0    # נפח מסחר פי X מהממוצע = Volume Spike
RS_BENCHMARK = 'SPY'     # מדד ייחוס לחישוב Relative Strength
DAILY_DROP_ALERT = -5.0  # ירידה של X% ביום = התראה דחופה

print("=" * 50)
print("   סורק מניות — NYSE + NASDAQ  (Cloud)")
print(f"   {datetime.now().strftime('%d/%m/%Y  %H:%M')}")
print("=" * 50)

# ─── פורטפוליו אישי — מחירי כניסה ─────────────────────
PORTFOLIO = [
    {'ticker': 'IBIT', 'entry': 38.72},
    {'ticker': 'BMNR', 'entry': 19.78},
    {'ticker': 'TSLA', 'entry': 343.22},
    {'ticker': 'VOO',  'entry': 623.64},
]

# ─── קבלת רשימת טיקרים ───────────────────────────────
def _clean_ticker(sym):
    sym = sym.strip().upper()
    if not sym or not sym.replace('.', '').replace('-', '').isalpha():
        return None
    if len(sym) > 5:
        return None
    return sym

def get_all_tickers():
    """
    מנסה כמה מקורות לפי סדר:
    1. GitHub mirror של רשימת NASDAQ/NYSE (rreichel3/US-Stock-Symbols)
    2. NASDAQ FTP הרשמי (חסום ברנרי GitHub Actions)
    3. Wikipedia - S&P 500 (גיבוי)
    4. רשימה מובנית מצומצמת
    """
    headers = {'User-Agent': 'Mozilla/5.0'}
    tickers = set()

    # מקור 1 — GitHub mirror (rreichel3/US-Stock-Symbols)
    mirror_urls = [
        "https://raw.githubusercontent.com/rreichel3/US-Stock-Symbols/main/nasdaq/nasdaq_tickers.txt",
        "https://raw.githubusercontent.com/rreichel3/US-Stock-Symbols/main/nyse/nyse_tickers.txt",
    ]
    for url in mirror_urls:
        try:
            r = requests.get(url, headers=headers, timeout=30)
            if r.status_code == 200 and r.text.strip():
                for line in r.text.strip().split('\n'):
                    s = _clean_ticker(line)
                    if s:
                        tickers.add(s)
                print(f"  ✓ נטען: {url.split('/')[-1]}")
        except Exception as e:
            print(f"  ✗ מקור 1 נכשל: {e}")

    if len(tickers) >= 100:
        return sorted(tickers)

    # מקור 2 — NASDAQ FTP הרשמי
    for url in [
        "https://ftp.nasdaqtrader.com/SymbolDirectory/nasdaqlisted.txt",
        "https://ftp.nasdaqtrader.com/SymbolDirectory/otherlisted.txt"
    ]:
        try:
            r = requests.get(url, headers=headers, timeout=15)
            lines = r.text.strip().split('\n')
            for line in lines[1:-1]:
                parts = line.split('|')
                if len(parts) >= 7 and parts[6] == 'N':
                    s = _clean_ticker(parts[0])
                    if s:
                        tickers.add(s)
        except Exception as e:
            print(f"  ✗ מקור 2 נכשל: {e}")

    if len(tickers) >= 100:
        return sorted(tickers)

    # מקור 3 — Wikipedia S&P 500
    try:
        wiki_url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
        tables = pd.read_html(wiki_url)
        if tables:
            for sym in tables[0]['Symbol'].astype(str).tolist():
                s = _clean_ticker(sym.replace('.', '-'))
                if s:
                    tickers.add(s)
            print(f"  ✓ נטען מ-Wikipedia: {len(tickers)} טיקרים")
    except Exception as e:
        print(f"  ✗ מקור 3 נכשל: {e}")

    if len(tickers) >= 100:
        return sorted(tickers)

    # מקור 4 — רשימה מובנית (גיבוי אחרון)
    print("  ⚠️  שימוש ברשימה מובנית כגיבוי")
    fallback = [
        "AAPL","MSFT","GOOGL","GOOG","AMZN","NVDA","META","TSLA","BRK-B","LLY",
        "AVGO","JPM","V","WMT","XOM","UNH","MA","PG","JNJ","HD","COST","ORCL",
        "ABBV","BAC","NFLX","KO","CRM","AMD","PEP","MRK","CVX","ADBE","TMO",
        "LIN","ACN","CSCO","MCD","ABT","WFC","IBM","TXN","NOW","GE","DHR",
        "AXP","PM","DIS","CAT","ISRG","INTU","VZ","AMGN","MS","GS","UNP",
        "PFE","RTX","NEE","SPGI","QCOM","LOW","T","PGR","UBER","BKNG","HON",
        "SYK","BLK","TJX","C","SCHW","BX","BSX","VRTX","ANET","MMC","LMT",
        "PLD","ADP","REGN","DE","ELV","MDLZ","ETN","ADI","FI","PANW","CB",
        "MU","BMY","KLAC","ICE","AMAT","SBUX","PYPL","SHOP","SNOW","PLTR",
        "COIN","MSTR","CRWD","ABNB","SQ","ROKU","ZM","DOCU","NET","DDOG",
        "MDB","TEAM","OKTA","TWLO","ZS","FTNT","CHWY","TDOC","PINS","SNAP",
        "LYFT","DASH","RBLX","U","HOOD","SOFI","AFRM","UPST","NU","LCID",
        "RIVN","NIO","XPEV","LI","BYD","F","GM","STLA","TM","HMC",
    ]
    tickers.update(fallback)
    return sorted(tickers)

# ─── חישוב ATR ───────────────────────────────────────
def calculate_atr(df, period=14):
    h, l, c = df['High'], df['Low'], df['Close']
    pc = c.shift(1)
    tr = pd.concat([(h - l), (h - pc).abs(), (l - pc).abs()], axis=1).max(axis=1)
    return tr.ewm(span=period, adjust=False).mean()

# ─── סריקה ─────────────────────────────────────────
def scan_batch(tickers, spy_return_1m=None):
    passing = []
    batch_size = 150
    total_batches = (len(tickers) + batch_size - 1) // batch_size
    for i in range(0, len(tickers), batch_size):
        batch = tickers[i:i+batch_size]
        bn = i // batch_size + 1
        print(f"  Batch {bn}/{total_batches}  ({len(batch)} טיקרים)...")
        try:
            raw = yf.download(
                ' '.join(batch), period='1y', group_by='ticker',
                auto_adjust=True, progress=False, threads=True, timeout=60
            )
        except Exception as e:
            print(f"  שגיאת הורדה: {e}")
            time.sleep(2)
            continue
        for t in batch:
            try:
                d = raw[t] if len(batch) > 1 else raw
                if d.empty or len(d) < 155:
                    continue
                d = d.dropna()
                if len(d) < 155:
                    continue
                price  = float(d['Close'].iloc[-1])
                ma150  = float(d['Close'].rolling(150).mean().iloc[-1])
                atr    = float(calculate_atr(d).iloc[-1])
                vol    = float(d['Volume'].iloc[-1])
                if any(np.isnan([price, ma150, atr])):
                    continue
                if price <= ma150 or price < 1.0:
                    continue
                if not (3.5 <= atr <= 7.0):
                    continue

                # Volume Spike — נפח יחסית לממוצע של 50 ימים
                vol_avg50 = float(d['Volume'].tail(50).mean())
                vol_ratio = vol / vol_avg50 if vol_avg50 > 0 else 0

                # Relative Strength — תשואה של המניה מול SPY ב-21 ימי מסחר (חודש)
                if len(d) >= 22:
                    stock_ret_1m = (price / float(d['Close'].iloc[-22]) - 1) * 100
                else:
                    stock_ret_1m = 0
                rs_score = None
                if spy_return_1m is not None and spy_return_1m != 0:
                    rs_score = round(stock_ret_1m / spy_return_1m, 2)

                passing.append({
                    'Ticker': t,
                    'Price': round(price, 2),
                    'MA150': round(ma150, 2),
                    'vs_MA_%': round((price / ma150 - 1) * 100, 1),
                    'ATR14': round(atr, 2),
                    'Volume': int(vol),
                    'Vol_Ratio': round(vol_ratio, 2),
                    'Ret_1M_%': round(stock_ret_1m, 1),
                    'RS': rs_score,
                })
            except Exception:
                continue
        time.sleep(0.3)
    return passing

# ─── SPY Benchmark (לחישוב Relative Strength) ────────
def get_spy_return_1m():
    try:
        d = yf.Ticker(RS_BENCHMARK).history(period='2mo', auto_adjust=True).dropna()
        if len(d) >= 22:
            return (float(d['Close'].iloc[-1]) / float(d['Close'].iloc[-22]) - 1) * 100
    except Exception as e:
        print(f"  שגיאה בחישוב SPY: {e}")
    return None

# ─── אימות שווי שוק ──────────────────────────────────
def add_market_cap(stocks):
    final = []
    print(f"\nמאמת שווי שוק עבור {len(stocks)} מניות...")
    for i, s in enumerate(stocks):
        try:
            fi = yf.Ticker(s['Ticker']).fast_info
            mc = getattr(fi, 'market_cap', None)
            if mc and mc >= 100_000_000:
                s['Market_Cap_M'] = round(mc / 1_000_000, 1)
                final.append(s)
        except Exception:
            continue
        if i % 20 == 0 and i > 0:
            time.sleep(0.5)
    return final

def enrich_top_companies(stocks):
    """מוסיף שם חברה רק ל-TOP (לאחר שחתכנו ל-20)."""
    for s in stocks:
        try:
            info = yf.Ticker(s['Ticker']).info
            s['Company'] = info.get('shortName') or info.get('longName') or s['Ticker']
        except Exception:
            s['Company'] = s['Ticker']
    return stocks

# ─── ניתוח פורטפוליו אישי ────────────────────────────
def analyze_portfolio(holdings):
    results = []
    for h in holdings:
        t = h['ticker']
        entry = h['entry']
        try:
            d = yf.Ticker(t).history(period='1y', auto_adjust=True)
            if d.empty:
                results.append({
                    'Ticker': t, 'Entry': entry, 'Price': None,
                    'Gain_$': None, 'Gain_%': None,
                    'MA150': None, 'vs_MA_%': None, 'ATR14': None,
                    'Day_Change_%': None,
                    'Status': 'אין נתונים', 'Alerts': []
                })
                continue
            d = d.dropna()
            price = float(d['Close'].iloc[-1])
            prev_close = float(d['Close'].iloc[-2]) if len(d) >= 2 else price
            ma150 = float(d['Close'].rolling(150).mean().iloc[-1]) if len(d) >= 150 else None
            atr   = float(calculate_atr(d).iloc[-1])   if len(d) >= 15  else None
            gain_pct = (price / entry - 1) * 100
            gain_abs = price - entry
            day_change = (price / prev_close - 1) * 100 if prev_close > 0 else 0
            flags, alerts = [], []
            if ma150 is not None:
                if price > ma150:
                    flags.append('✅ מעל MA150')
                else:
                    flags.append('⚠️ מתחת MA150')
                    alerts.append(f'{t} שבר את ה-MA150 (${price:.2f} מתחת ל-${ma150:.2f})')
            if atr is not None:
                if 3.5 <= atr <= 7.0:
                    flags.append('✅ ATR בטווח')
                elif atr < 3.5:
                    flags.append('ℹ️ ATR נמוך')
                else:
                    flags.append('ℹ️ ATR גבוה')
            if day_change <= DAILY_DROP_ALERT:
                alerts.append(f'{t} ירדה {day_change:+.1f}% היום')
            results.append({
                'Ticker': t,
                'Entry':  round(entry, 2),
                'Price':  round(price, 2),
                'Gain_$': round(gain_abs, 2),
                'Gain_%': round(gain_pct, 1),
                'MA150':  round(ma150, 2) if ma150 is not None else None,
                'vs_MA_%': round((price / ma150 - 1) * 100, 1) if ma150 is not None else None,
                'ATR14':  round(atr, 2) if atr is not None else None,
                'Day_Change_%': round(day_change, 2),
                'Status': ' | '.join(flags) if flags else '-',
                'Alerts': alerts
            })
        except Exception as e:
            results.append({
                'Ticker': t, 'Entry': entry, 'Price': None,
                'Gain_$': None, 'Gain_%': None,
                'MA150': None, 'vs_MA_%': None, 'ATR14': None,
                'Day_Change_%': None,
                'Status': f'שגיאה: {str(e)[:40]}', 'Alerts': []
            })
    return results

# ─── ניהול היסטוריה — Hot Picks ──────────────────────
def load_history(output_dir):
    """טוען היסטוריה מהריפו. הקובץ יושב בשורש הריפו."""
    repo_root = os.environ.get('GITHUB_WORKSPACE', os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(repo_root, HISTORY_FILE)
    if not os.path.exists(path):
        return {'scans': []}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {'scans': []}

def save_history(history, output_dir):
    """שומר היסטוריה ב-history.json בשורש הריפו."""
    repo_root = os.environ.get('GITHUB_WORKSPACE', os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(repo_root, HISTORY_FILE)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)
    return path

def update_history(history, today_tickers):
    """מוסיף את היום להיסטוריה (אחד ליום, גם אם רצים כמה פעמים)."""
    today = datetime.utcnow().strftime('%Y-%m-%d')
    # הסר רישום קיים של היום
    history['scans'] = [s for s in history.get('scans', []) if s['date'] != today]
    history['scans'].append({'date': today, 'tickers': sorted(set(today_tickers))})
    # שמור רק X ימים אחרונים
    history['scans'] = sorted(history['scans'], key=lambda s: s['date'])[-HISTORY_KEEP_DAYS:]
    return history

def calc_hot_picks(history):
    """לכל טיקר — סופר כמה ימים *רצופים* הוא הופיע (כולל היום)."""
    scans = sorted(history.get('scans', []), key=lambda s: s['date'], reverse=True)
    if not scans:
        return {}
    streaks = {}
    today_tickers = set(scans[0]['tickers'])
    for t in today_tickers:
        streak = 1
        for prev in scans[1:]:
            if t in set(prev['tickers']):
                streak += 1
            else:
                break
        streaks[t] = streak
    return streaks

# ─── שמירת Excel ─────────────────────────────────────
def save_excel(final, portfolio, output_dir, ts):
    from openpyxl.styles import PatternFill, Font, Alignment
    cols = ['Ticker', 'Company', 'Hot', 'Price', 'MA150', 'vs_MA_%', 'ATR14',
            'Vol_Ratio', 'Ret_1M_%', 'RS', 'Market_Cap_M', 'Volume']
    df = pd.DataFrame(final) if final else pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = ''
    df = df[cols]

    port_cols = ['Ticker', 'Entry', 'Price', 'Day_Change_%', 'Gain_$', 'Gain_%',
                 'MA150', 'vs_MA_%', 'ATR14', 'Status']
    df_port = pd.DataFrame(portfolio) if portfolio else pd.DataFrame(columns=port_cols)
    for c in port_cols:
        if c not in df_port.columns:
            df_port[c] = ''
    df_port = df_port[port_cols]

    path = os.path.join(output_dir, f"stock_scan_{ts}.xlsx")
    hfill = PatternFill("solid", fgColor="1F4E79")
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        # גיליון 1 — פורטפוליו
        df_port.to_excel(writer, index=False, sheet_name='Portfolio')
        # גיליון 2 — תוצאות הסריקה
        df.to_excel(writer, index=False, sheet_name='Stocks')
        for sheet_name in ('Portfolio', 'Stocks'):
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = hfill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                try:
                    mx = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 30)
                except Exception:
                    pass
        if df.empty:
            writer.sheets['Stocks'].cell(row=2, column=1, value='No stocks matched criteria')
    return path

# ─── שליחת מייל דרך SMTP (Gmail) ─────────────────────
def send_email_smtp(to_addr, subject, html_body, attachment_path=None):
    """
    דורש משתני סביבה:
      SMTP_USER     — כתובת ה-Gmail השולחת
      SMTP_PASSWORD — App Password (סיסמה לאפליקציה) של Gmail
    """
    smtp_user = os.environ.get('SMTP_USER')
    smtp_pass = os.environ.get('SMTP_PASSWORD')
    if not smtp_user or not smtp_pass:
        raise RuntimeError("חסרים משתני סביבה SMTP_USER / SMTP_PASSWORD")

    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject
    msg['From'] = smtp_user
    msg['To'] = to_addr

    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(html_body, 'html', 'utf-8'))
    msg.attach(alt)

    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, 'rb') as f:
            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename="{os.path.basename(attachment_path)}"'
        )
        msg.attach(part)

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as server:
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, [to_addr], msg.as_string())

# ─── בניית HTML לאימייל ──────────────────────────────
def build_portfolio_html(portfolio):
    if not portfolio:
        return ""
    rows = ""
    for p in portfolio:
        price = p.get('Price')
        gain_pct = p.get('Gain_%')
        gain_abs = p.get('Gain_$')
        ma150 = p.get('MA150')
        vs_ma = p.get('vs_MA_%')
        atr = p.get('ATR14')
        status = p.get('Status', '-')
        if price is None:
            rows += f"""<tr>
              <td style="padding:6px 10px;border:1px solid #ddd;font-weight:bold;color:#1F4E79">{p['Ticker']}</td>
              <td style="padding:6px 10px;border:1px solid #ddd;text-align:right">${p['Entry']:.2f}</td>
              <td colspan="6" style="padding:6px 10px;border:1px solid #ddd;text-align:center;color:#999">{status}</td>
            </tr>"""
            continue
        gcol = '#16a34a' if gain_pct >= 0 else '#dc2626'
        mcol = '#16a34a' if (vs_ma or 0) > 0 else '#dc2626'
        ma150_str = f"${ma150:.2f}" if ma150 is not None else "-"
        vs_ma_str = f"{vs_ma:+.1f}%" if vs_ma is not None else "-"
        rows += f"""<tr>
          <td style="padding:6px 10px;border:1px solid #ddd;font-weight:bold;color:#1F4E79">{p['Ticker']}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right">${p['Entry']:.2f}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right;font-weight:bold">${price:.2f}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right;color:{gcol}">{gain_abs:+.2f}$</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right;color:{gcol};font-weight:bold">{gain_pct:+.1f}%</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right">{ma150_str}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right;color:{mcol}">{vs_ma_str}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;font-size:12px">{status}</td>
        </tr>"""
    return f"""
  <div style="padding:20px;background:#fffbeb;border-top:3px solid #f59e0b">
    <h3 style="color:#92400e;margin-top:0">💼 התיק שלי</h3>
    <table style="width:100%;border-collapse:collapse;font-size:13px" dir="ltr">
      <thead>
        <tr style="background:#92400e;color:white">
          <th style="padding:8px 10px;text-align:left">Ticker</th>
          <th style="padding:8px 10px;text-align:right">Entry</th>
          <th style="padding:8px 10px;text-align:right">Price</th>
          <th style="padding:8px 10px;text-align:right">P/L $</th>
          <th style="padding:8px 10px;text-align:right">P/L %</th>
          <th style="padding:8px 10px;text-align:right">MA150</th>
          <th style="padding:8px 10px;text-align:right">vs MA</th>
          <th style="padding:8px 10px;text-align:left">Status</th>
        </tr>
      </thead>
      <tbody>{rows}</tbody>
    </table>
  </div>"""

def build_urgent_alert_html(alert_items, now):
    items_html = "".join(f'<li style="padding:8px 0;font-size:15px">{a}</li>' for a in alert_items)
    return f"""
<div style="font-family:Arial,sans-serif;max-width:700px;margin:0 auto" dir="rtl">
  <div style="background:#dc2626;color:white;padding:20px;border-radius:8px 8px 0 0">
    <h2 style="margin:0">⚠️ התראה דחופה — תיק אישי</h2>
    <p style="margin:5px 0 0 0;opacity:0.9">{now.strftime('%A, %d/%m/%Y | %H:%M UTC')}</p>
  </div>
  <div style="padding:20px;background:white;border:2px solid #dc2626;border-top:none;border-radius:0 0 8px 8px">
    <ul style="margin:0;padding-right:20px;color:#991b1b">{items_html}</ul>
    <p style="margin-top:20px;color:#666;font-size:13px">
      💡 <b>מה לעשות?</b> בדוק את המניות, שקול עדכון Stop Loss או יציאה לפי האסטרטגיה שלך.
    </p>
  </div>
</div>"""

def build_email_html(final, tickers, now, portfolio=None, hot_streaks=None):
    rows_html = ""
    for s in final:
        color = 'green' if s['vs_MA_%'] > 0 else 'red'
        streak = (hot_streaks or {}).get(s['Ticker'], 0)
        hot_badge = f'<span style="background:#f59e0b;color:white;padding:2px 6px;border-radius:4px;font-size:11px">🔥 {streak}d</span>' if streak >= HOT_PICK_DAYS else ''
        vol_ratio = s.get('Vol_Ratio') or 0
        vol_badge = f'<span style="background:#16a34a;color:white;padding:2px 6px;border-radius:4px;font-size:11px">📈 {vol_ratio:.1f}x</span>' if vol_ratio >= VOL_SPIKE_RATIO else f'{vol_ratio:.1f}x'
        rs = s.get('RS')
        rs_str = f'{rs:.2f}' if rs is not None else '-'
        rs_color = '#16a34a' if rs and rs > 1 else '#666'
        ret_1m = s.get('Ret_1M_%', 0)
        ret_color = '#16a34a' if ret_1m > 0 else '#dc2626'
        rows_html += f"""<tr>
          <td style="padding:6px 10px;border:1px solid #ddd;font-weight:bold;color:#1F4E79">{s['Ticker']} {hot_badge}</td>
          <td style="padding:6px 10px;border:1px solid #ddd">{s.get('Company','')[:30]}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right">${s['Price']:.2f}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right">${s['MA150']:.2f}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right;color:{color}">{s['vs_MA_%']:+.1f}%</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right;font-weight:bold">{s['ATR14']:.2f}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right">{vol_badge}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right;color:{ret_color}">{ret_1m:+.1f}%</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right;color:{rs_color};font-weight:bold">{rs_str}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:right">${s.get('Market_Cap_M',0):,.0f}M</td>
        </tr>"""

    no_results = '<tr><td colspan="10" style="text-align:center;padding:20px;color:#999">לא נמצאו מניות</td></tr>'
    return f"""
<div style="font-family:Arial,sans-serif;max-width:1100px;margin:0 auto" dir="rtl">
  <div style="background:#1F4E79;color:white;padding:20px;border-radius:8px 8px 0 0">
    <h2 style="margin:0">📊 סריקת מניות — TOP {MAX_RESULTS} לפי שווי שוק</h2>
    <p style="margin:5px 0 0 0;opacity:0.85">{now.strftime('%A, %d/%m/%Y | %H:%M UTC')}</p>
  </div>
  <div style="background:#f0f4f8;padding:15px;display:flex;gap:20px">
    <div style="background:white;padding:12px 20px;border-radius:6px;text-align:center;flex:1">
      <div style="font-size:28px;font-weight:bold;color:#1F4E79">{len(tickers):,}</div>
      <div style="color:#666;font-size:13px">מניות נסרקו</div>
    </div>
    <div style="background:white;padding:12px 20px;border-radius:6px;text-align:center;flex:1">
      <div style="font-size:28px;font-weight:bold;color:#16a34a">{len(final)}</div>
      <div style="color:#666;font-size:13px">מוצגות (TOP לפי שווי שוק)</div>
    </div>
    <div style="background:white;padding:12px 20px;border-radius:6px;text-align:center;flex:1">
      <div style="font-size:28px;font-weight:bold;color:#d97706">ATR 3.5–7.0</div>
      <div style="color:#666;font-size:13px">טווח</div>
    </div>
  </div>
  {build_portfolio_html(portfolio) if portfolio else ''}
  <div style="padding:20px;background:white">
    <h3 style="color:#1F4E79;margin-top:0">מניות מובילות (ממוינות לפי שווי שוק)</h3>
    <p style="color:#666;font-size:12px;margin:0 0 10px 0">🔥 = Hot Pick (נמצא בסריקה {HOT_PICK_DAYS}+ ימים ברצף) | 📈 = Volume Spike (×{VOL_SPIKE_RATIO:g}) | RS &gt; 1 = חזקה מ-SPY</p>
    <table style="width:100%;border-collapse:collapse;font-size:13px" dir="ltr">
      <thead>
        <tr style="background:#1F4E79;color:white">
          <th style="padding:8px 10px;text-align:left">Ticker</th>
          <th style="padding:8px 10px;text-align:left">Company</th>
          <th style="padding:8px 10px;text-align:right">Price</th>
          <th style="padding:8px 10px;text-align:right">MA150</th>
          <th style="padding:8px 10px;text-align:right">% vs MA</th>
          <th style="padding:8px 10px;text-align:right">ATR14</th>
          <th style="padding:8px 10px;text-align:right">Vol Ratio</th>
          <th style="padding:8px 10px;text-align:right">Ret 1M</th>
          <th style="padding:8px 10px;text-align:right">RS</th>
          <th style="padding:8px 10px;text-align:right">Market Cap</th>
        </tr>
      </thead>
      <tbody>{rows_html if rows_html else no_results}</tbody>
    </table>
  </div>
  <div style="background:#f8f8f8;padding:12px 20px;border-radius:0 0 8px 8px;font-size:12px;color:#999;text-align:center">
    נוצר אוטומטית על ידי סורק המניות | NYSE + NASDAQ | GitHub Actions
  </div>
</div>"""

# ─── תוכנית ראשית ─────────────────────────────────────
if __name__ == "__main__":
    OUTPUT_DIR = os.environ.get('OUTPUT_DIR', os.path.dirname(os.path.abspath(__file__)))
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    now = datetime.now()
    ts = now.strftime('%Y-%m-%d_%H%M')

    print("\n📥 מוריד רשימת טיקרים...")
    tickers = get_all_tickers()
    print(f"✅ סה\"כ טיקרים: {len(tickers):,}")

    print("\n💼 מנתח פורטפוליו אישי...")
    portfolio = analyze_portfolio(PORTFOLIO)
    for p in portfolio:
        print(f"  {p['Ticker']:6} Entry ${p['Entry']:.2f} → Price ${p.get('Price') or 0:.2f}  ({p.get('Gain_%') or 0:+.1f}%)")

    # ─── התראות דחופות לתיק ──────────────
    all_alerts = []
    for p in portfolio:
        all_alerts.extend(p.get('Alerts', []))
    to_addr = os.environ.get('MAIL_TO', 'helelit22@gmail.com')
    if all_alerts:
        print(f"\n⚠️  נמצאו {len(all_alerts)} התראות דחופות — שולח מייל נפרד...")
        try:
            urgent_subject = f"⚠️ התראה דחופה — {len(all_alerts)} אירועים בתיק | {now.strftime('%d/%m/%Y')}"
            urgent_html = build_urgent_alert_html(all_alerts, now)
            send_email_smtp(to_addr, urgent_subject, urgent_html)
            print("✅ מייל התראה דחופה נשלח")
        except Exception as e:
            print(f"⚠️ שגיאה במייל התראה: {e}")

    print(f"\n📈 מחשב SPY benchmark לחישוב Relative Strength...")
    spy_ret = get_spy_return_1m()
    print(f"  SPY 1M return: {spy_ret:+.2f}%" if spy_ret else "  SPY: n/a")

    print("\n🔍 סורק מחיר / MA150 / ATR / Volume / RS...")
    preliminary = scan_batch(tickers, spy_return_1m=spy_ret)
    print(f"\n✅ עברו סינון מחיר/ATR: {len(preliminary)} מניות")

    final = add_market_cap(preliminary)
    # מיין לפי שווי שוק (גדול לקטן) וקח TOP MAX_RESULTS
    final.sort(key=lambda x: x.get('Market_Cap_M', 0), reverse=True)
    final = final[:MAX_RESULTS]
    print(f"✅ TOP {MAX_RESULTS} לפי שווי שוק: {len(final)} מניות")

    # מוסיף שמות חברות רק ל-TOP
    print("\n🏷️  מושך שמות חברות...")
    final = enrich_top_companies(final)

    # ─── עדכון היסטוריה + Hot Picks ──────
    print("\n📚 מעדכן היסטוריה ומחשב Hot Picks...")
    history = load_history(OUTPUT_DIR)
    history = update_history(history, [s['Ticker'] for s in final])
    save_history(history, OUTPUT_DIR)
    hot_streaks = calc_hot_picks(history)
    hot_count = sum(1 for v in hot_streaks.values() if v >= HOT_PICK_DAYS)
    print(f"✅ Hot Picks (>={HOT_PICK_DAYS} ימים ברצף): {hot_count}")
    # מוסיף את המידע למניות
    for s in final:
        s['Hot'] = f"🔥 {hot_streaks.get(s['Ticker'], 0)}d" if hot_streaks.get(s['Ticker'], 0) >= HOT_PICK_DAYS else ''

    print("\n💾 שומר Excel...")
    xlsx_path = save_excel(final, portfolio, OUTPUT_DIR, ts)
    print(f"✅ נשמר: {xlsx_path}")

    print("\n📧 שולח מייל ראשי...")
    subject = f"📊 סריקת מניות — TOP {MAX_RESULTS} | {now.strftime('%d/%m/%Y %H:%M')}"
    html = build_email_html(final, tickers, now, portfolio=portfolio, hot_streaks=hot_streaks)
    try:
        send_email_smtp(to_addr, subject, html, attachment_path=xlsx_path)
        print(f"✅ מייל נשלח בהצלחה ל-{to_addr}")
    except Exception as e:
        print(f"⚠️  שגיאה בשליחת מייל: {e}")
        sys.exit(1)

    print("\n" + "=" * 50)
    print(f"   ✅ סריקה הושלמה — {len(final)} מניות (TOP {MAX_RESULTS})")
    print(f"   🔥 Hot Picks: {hot_count} | ⚠️ התראות תיק: {len(all_alerts)}")
    print("=" * 50)
